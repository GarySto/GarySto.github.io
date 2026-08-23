#!/usr/bin/env python3
"""
build_articles.py

Reads gary-hq/logs/Performance_Log.xlsx (the LinkedIn content + performance
log — NOT opportunity_log.csv, which tracks interview/speaking outcomes,
a different thing entirely) and turns the best eligible unused post into a
real article page on the site.

What it does NOT do: push straight to main. It writes the new files into
a checkout of GarySto.github.io (passed via --site-dir); the workflow that
calls this script then opens a Pull Request. Merging that PR is the
single yes/no approval step — publishing to a public site always needs
one explicit action from Gary, even when the text itself needs no edits.

Eligibility filters (all must pass):
  - Post type is "Post"
  - Pillar is NOT "3 - Lessons & the human stuff" (personal content is
    skipped by design — this pipeline is for subject-relevant posts only)
  - Full post text is at least MIN_WORDS words (skips short/low-effort posts)
  - Not already used, tracked in gary-hq/logs/site_published.json

Ranking: highest Impressions among what's left. If nothing recent
qualifies, this naturally works back through the whole backlog, since it
ranks the full eligible set, not just recent rows.
"""
import argparse
import json
import re
from datetime import date
from pathlib import Path

import openpyxl

MIN_WORDS = 100
EXCLUDED_PILLAR_PREFIX = "3"  # "3 - Lessons & the human stuff"

ARTICLE_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title} — Gary Stow</title>
<meta name="description" content="{description}">
<meta name="robots" content="index, follow">
<link rel="canonical" href="https://garystow.co.uk/articles/{slug}.html">
<script type="application/ld+json">
{{
  "@context": "https://schema.org",
  "@type": "Article",
  "headline": "{title}",
  "author": {{"@type": "Person", "name": "Gary Stow", "url": "https://garystow.co.uk/"}},
  "datePublished": "{iso_date}",
  "url": "https://garystow.co.uk/articles/{slug}.html"
}}
</script>
<!-- Google tag (gtag.js) -->
<script async src="https://www.googletagmanager.com/gtag/js?id=G-5S939E9N5T"></script>
<script>
  window.dataLayer = window.dataLayer || [];
  function gtag(){{dataLayer.push(arguments);}}
  gtag('js', new Date());
  gtag('config', 'G-5S939E9N5T');
</script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Mono:wght@400;500&family=DM+Sans:ital,wght@0,300;0,400;0,500;1,300&display=swap" rel="stylesheet">
<style>
  :root {{ --ink:#0f0f0f; --paper:#f5f2eb; --accent:#c8410a; --muted:#6b6560; --rule:#d4cfc7;
    --mono:'DM Mono',monospace; --serif:'DM Serif Display',serif; --sans:'DM Sans',sans-serif; }}
  *,*::before,*::after {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ background:var(--paper); color:var(--ink); font-family:var(--sans); font-size:17px; line-height:1.75; font-weight:300; }}
  header {{ border-bottom:1px solid var(--ink); padding:2rem 4rem; display:flex; justify-content:space-between; align-items:baseline; }}
  .logo, .back {{ font-family:var(--mono); font-size:0.8rem; letter-spacing:0.08em; text-transform:uppercase; color:var(--muted); text-decoration:none; }}
  .back:hover {{ color:var(--accent); }}
  main {{ max-width:720px; margin:0 auto; padding:5rem 2rem; }}
  .meta {{ font-family:var(--mono); font-size:0.75rem; color:var(--muted); text-transform:uppercase; letter-spacing:0.05em; margin-bottom:1rem; }}
  h1 {{ font-family:var(--serif); font-size:clamp(2.2rem,5vw,3.2rem); font-weight:400; line-height:1.15; margin-bottom:2rem; }}
  .body p {{ margin-bottom:1.4rem; color:var(--ink); }}
  .origin {{ margin-top:3rem; padding-top:1.5rem; border-top:1px solid var(--rule); font-family:var(--mono); font-size:0.8rem; color:var(--muted); }}
  .origin a {{ color:var(--accent); }}
</style>
</head>
<body>
<header>
  <a href="../" class="logo">Gary Stow</a>
  <a href="../articles/" class="back">← All articles</a>
</header>
<main>
  <div class="meta">{pillar}</div>
  <h1>{title}</h1>
  <div class="body">
{body_html}
  </div>
  <div class="origin">Originally posted on LinkedIn{post_url_html}.</div>
</main>
</body>
</html>
"""


def slugify(text: str) -> str:
    text = (text or "post").lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text[:60] or "post"


def load_json_set(path: Path) -> set:
    if path.exists():
        return set(json.loads(path.read_text()))
    return set()


def save_json_set(path: Path, values: set) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(sorted(values), indent=2))


def to_html_paragraphs(text: str) -> str:
    paras = [p.strip() for p in (text or "").split("\n") if p.strip()]
    return "\n".join(f"    <p>{p}</p>" for p in paras)


def pick_next_post(rows, used: set):
    candidates = []
    for r in rows:
        text = r.get("Full post text") or ""
        pillar = str(r.get("Pillar") or "")
        key = r.get("Post URL") or f"no-url:{slugify(r.get('Topic'))}"
        if r.get("Post type") != "Post":
            continue
        if pillar.strip().startswith(EXCLUDED_PILLAR_PREFIX):
            continue
        if len(text.split()) < MIN_WORDS:
            continue
        if key in used:
            continue
        candidates.append(r)

    if not candidates:
        return None

    candidates.sort(key=lambda r: (r.get("Impressions") or 0), reverse=True)
    return candidates[0]


def write_article(site_dir: Path, pick: dict) -> str:
    title = pick.get("Topic") or "Untitled"
    slug = slugify(title)
    body_html = to_html_paragraphs(pick.get("Full post text"))
    post_url = pick.get("Post URL") or ""
    post_url_html = f' (<a href="{post_url}">original post</a>)' if post_url else ""
    description = (pick.get("Hook (first line)") or title)[:155]

    html = ARTICLE_TEMPLATE.format(
        title=title,
        slug=slug,
        description=description,
        pillar=pick.get("Pillar") or "",
        body_html=body_html,
        post_url_html=post_url_html,
        iso_date=date.today().isoformat(),
    )

    articles_dir = site_dir / "articles"
    articles_dir.mkdir(parents=True, exist_ok=True)
    (articles_dir / f"{slug}.html").write_text(html)
    return slug


def update_articles_index(site_dir: Path, slug: str, title: str) -> None:
    index_path = site_dir / "articles" / "index.html"
    if not index_path.exists():
        return
    content = index_path.read_text()
    entry = (
        f'    <li class="article-item"><a href="{slug}.html">{title}</a>'
        f'<span class="article-meta">{date.today().isoformat()}</span></li>\n'
    )
    content = content.replace(
        '<li class="empty">Nothing published here yet — check back soon.</li>\n',
        "",
    )
    content = content.replace(
        "<!-- ARTICLE_LIST_START -->\n",
        "<!-- ARTICLE_LIST_START -->\n" + entry,
    )
    index_path.write_text(content)


def update_sitemap(site_dir: Path, slug: str) -> None:
    sitemap_path = site_dir / "sitemap.xml"
    if not sitemap_path.exists():
        return
    content = sitemap_path.read_text()
    url = f"https://garystow.co.uk/articles/{slug}.html"
    if url in content:
        return
    entry = f"  <url>\n    <loc>{url}</loc>\n  </url>\n"
    content = content.replace(
        "<!-- SITEMAP_URLS_START -->\n",
        "<!-- SITEMAP_URLS_START -->\n" + entry,
    )
    sitemap_path.write_text(content)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--log", default="logs/Performance_Log.xlsx")
    parser.add_argument("--published", default="logs/site_published.json")
    parser.add_argument("--site-dir", default="../site")
    args = parser.parse_args()

    log_path = Path(args.log)
    published_path = Path(args.published)
    site_dir = Path(args.site_dir)

    if not log_path.exists():
        print(f"No log found at {log_path} — nothing to do.")
        return

    wb = openpyxl.load_workbook(log_path, data_only=True)
    ws = wb["Performance Log"]
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

    used = load_json_set(published_path)
    pick = pick_next_post(rows, used)

    if pick is None:
        print("No eligible unused posts found (checked the whole backlog).")
        return

    slug = write_article(site_dir, pick)
    update_articles_index(site_dir, slug, pick.get("Topic") or "Untitled")
    update_sitemap(site_dir, slug)

    key = pick.get("Post URL") or f"no-url:{slug}"
    used.add(key)
    save_json_set(published_path, used)

    print(f"Article written: {site_dir}/articles/{slug}.html")


if __name__ == "__main__":
    main()
